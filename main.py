import requests
import openpyxl
from xlutils.view import View
from random import randint

client_id = "a3e059563d7fd3372b49b37f00a00bcf"
iterations = 10
follower_min = 10000

duplicate_count = 0
low_follower_count = 0
no_contact_count = 0
good_count = 0

class SoundCloudUser:
    def __init__(self, username=None, id=None):
        if id != None:
            self.id = id
        else:
            self.username = username
            api = ("http://api.soundcloud.com/resolve?url=https://soundcloud.com/" + str(self.username) + "&client_id=" + client_id)
            r = requests.get(api)
            try:
                self.id = r.json()['id']
            except:
                print("Invalid Initial Username")
        
    def follower_list(self):
        api = ("http://api.soundcloud.com/users/" + str(self.id) + "/followings?client_id=" + client_id)
        r =  requests.get(api)
        try:
            collection = r.json()['collection']
            return collection
        except:
            print("Invalid Response :: Follower List")
            
    def cleanDescription(self, desc):
        words = desc.split()
        contacts = ""
        for word in words:
            if "@" in word:
                print(word)
                contacts += " " + word 
        return contacts
    
    def print_data(self):
        return (self.username + " :: " + str(self.userId) + " :: " + self.contacts + " :: " + str(self.follow_count) + " :: " + self.url)
    
    def get_summary(self):
        return 'Name: {:>24}  Followers: {:>10}  Contacts: {:>3}'.format(self.username, str(self.follow_count), str(len(self.contacts.split())))
    
    def gather_data(self):
        api = ("http://api.soundcloud.com/users/" + str(self.id) + "?client_id=" + client_id)
        r = requests.get(api)    
        try:
            self.username = str(r.json()['username'])
            self.userId = r.json()['id']
            self.follow_count = r.json()['followers_count']
            self.url = str("https://soundcloud.com/" + str(r.json()['permalink']))

            desc = str(r.json()['description']).replace('\n', " ")
            words = desc.split()
            contacts = ""
            for word in words:
                if "@" in word and "." in word:
                    contacts += " " + word 
            contacts = str(contacts)
            self.contacts = contacts
        except:
            print("Invalid Response :: Print Data")
            

class SpreadSheet:
    def __init__(self, name, is_view):
        self.name = name
        self.is_view = is_view
        if is_view is not True:
            self.wb = openpyxl.Workbook()
        else:
            self.wb = openpyxl.load_workbook(self.name)
        self.sheet = self.wb.active
        
    def set_cell(self, row, col, val):
        if self.is_view is not True:
            self.sheet.cell(row = row, column = col).value = val
            
    def view_cell(self, row, col):
        return self.sheet[str(col) + str(row)]
    
    def get_first_empty_row(self):
        first_empty_row = 1
        for cell in self.sheet["A"]:
            if cell.value:
                first_empty_row += 1
            else:
                return first_empty_row
        return first_empty_row
    
    def id_exists(self, id):
        for cell in self.sheet["A"]:
            #print(str(cell.value) + " == " + str(id) + " --> " + str(cell.value == id))
            if str(cell.value) == str(id):
                return True
        return False
    
    def save(self):
        if self.is_view is not True:
            self.wb.save(self.name)
        
#main
user = SoundCloudUser("alex_shortt", None)
list = user.follower_list()
count = 0;

read = SpreadSheet("catalogue.xlsx", True)
write = SpreadSheet("add_to_catalogue.xlsx", False)

for x in range(0, iterations):
    for val in list:
        follow_user = SoundCloudUser(None, val['id'])
        follow_user.gather_data()
        
        read_duplicate = read.id_exists(follow_user.userId)
        write_duplicate = write.id_exists(follow_user.userId)
        
        if read_duplicate == True or write_duplicate == True:
            duplicate_count += 1
            print("Duplicate              --> " + follow_user.get_summary())
        elif follow_user.contacts == "": 
            no_contact_count += 1
            print("No Contacts            --> " + follow_user.get_summary())
        elif follow_user.follow_count < follower_min: 
            low_follower_count += 1
            print("Too Little Followers   --> " + follow_user.get_summary())
        else:
            good_count += 1
            print("Good                   --> " + follow_user.get_summary())
            empty_row = write.get_first_empty_row()
            write.set_cell(empty_row, 1, follow_user.userId)
            write.set_cell(empty_row, 2, follow_user.username)
            write.set_cell(empty_row, 3, follow_user.contacts)
            write.set_cell(empty_row, 4, follow_user.follow_count)
            write.set_cell(empty_row, 5, follow_user.url)
            
    length = len(list) - 1
    user = SoundCloudUser(None, list[randint(0, length)]['id'])
    list = user.follower_list()

print("")
print("Duplicates:    " + str(duplicate_count))
print("Low Followers: " + str(low_follower_count))
print("No Contacts:   " + str(no_contact_count))
print("Good:          " + str(good_count))

try:    
    write.save()
except:
    print("ERROR SAVING")
    print("    Make sure to delete add_to_catalogue.xlsx before running script")
    
    